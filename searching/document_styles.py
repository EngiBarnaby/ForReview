from openpyxl.styles import (
    PatternFill, Border, Side,
    Alignment, Font, GradientFill
)

# СТИЛЬ ШРИФТА
standard_font = Font(
    name='Calibri',
    size=9,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='000000'
)

title_font = Font(
    name='Calibri',
    size=9,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='000000'
)

# ЗАЛИВКА ЯЧЕЕК
white_fill = PatternFill(fill_type='solid', fgColor='FFFFFF')
iek_title_fill = PatternFill(fill_type='solid', fgColor='F6BA00')
iek_cell_fill = PatternFill(fill_type='solid', fgColor='FFF1C5')
comment_title_fill = PatternFill(fill_type='solid', fgColor='B8CCE4')
comment_cell_fill = PatternFill(fill_type='solid', fgColor='DCE6F1')
discount_title_fill = PatternFill(fill_type='solid', fgColor='FAC090')
discount_cell_fill = PatternFill(fill_type='solid', fgColor='FDEADA')

# ГРАНИЦЫ ЯЧЕЕК
bold_border = Border(
    left=Side(border_style='thick', color='000000'),
    right=Side(border_style='thick', color='000000'),
    top=Side(border_style='thick', color='000000'),
    bottom=Side(border_style='thick', color='000000'),
    diagonal=Side(border_style=None, color='000000'),
    diagonal_direction=0,
    outline=Side(border_style=None, color='000000'),
    vertical=Side(border_style=None, color='000000'),
    horizontal=Side(border_style=None, color='000000')
)

standard_border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
    diagonal=Side(border_style=None, color='000000'),
    diagonal_direction=0,
    outline=Side(border_style=None, color='000000'),
    vertical=Side(border_style=None, color='000000'),
    horizontal=Side(border_style=None, color='000000')
)

separate_left_bold_border = Border(
    left=Side(border_style='thick', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
    diagonal=Side(border_style=None, color='000000'),
    diagonal_direction=0,
    outline=Side(border_style=None, color='000000'),
    vertical=Side(border_style=None, color='000000'),
    horizontal=Side(border_style=None, color='000000')
)

# ВЫРАВНИВАНИЕ В ЯЧЕЙКАХ
alignment_center = Alignment(
    horizontal='center',
    vertical='top',
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=True,
    indent=0
)

alignment_left = Alignment(
    horizontal='left',
    vertical='top',
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=True,
    indent=0
)
