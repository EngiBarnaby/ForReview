import pandas as pd
import collections
import os
import threading

from asgiref.sync import async_to_sync
from application_info.services import LogService
from channels.layers import get_channel_layer

from rest_framework.utils import json, encoders

from django.core.files import File
from django.db import transaction, IntegrityError
from django.db.models import Count, Q, F, Value, IntegerField
from django.conf import settings
from django.http.response import FileResponse
from django.utils import timezone
from django.utils.encoding import escape_uri_path

from . import document_styles
from .exceptions import Interrupt
from .exceptions import InvalidUploadingData
from .exceptions import MultiEquipCategoryDownload
from .models import SearchProcess
from .models import EquipmentCategory
from .models import OurEquipment
from .models import OurEquipmentProperty
from .models import Competitor
from .models import CompetitorsEquipment
from .models import CompetitorsEquipmentProperty
from .models import KeyWord
from .models import EquipmentCategoryPropertyValue
from .models import EquipmentCategoryProperty
from .models import OurEquipmentImage

from urllib.request import urlretrieve, urlcleanup
from urllib.error import URLError

from openpyxl import Workbook
from openpyxl.drawing.image import Image


class SearchService:

    @staticmethod
    def check_interrupt(prc):
        prc.refresh_from_db()
        if prc.is_interrupted:
            raise Interrupt

    @staticmethod
    def remtime_str(sec: int):
        """
        Возвращает псевдо-оставшееся время до окончания пересчета (remtime - remaining time)
        в виде строки "1ч. 2м. 3с."

        sec <-  Кол-во оставшихся секунд до завершения пересчета
        """
        s = sec % (24 * 3600)
        h = sec // 3600
        s %= 3600
        m = sec // 60
        s %= 60
        return "%02dч. %02dм. %02dс." % (h, m, s)

    @staticmethod
    def process_log(proc: SearchProcess, log: str, rt: str):
        """
        Отправляет переданную строку состояния процесса в сокет соответствующего процесса.

        proc        <-  Объект процесса пересчета
        log         <-  Строка лога
        rem_count   <-  Строка лога
        """
        cl = get_channel_layer()
        message = json.dumps({'log': log, 'rt': rt}, cls=encoders.JSONEncoder, ensure_ascii=False)
        async_to_sync(cl.group_send)(f"search-progress-{proc.user_id}", {"type": "search_progress", "message": message})

    @staticmethod
    def get_prepare_data(f):
        """
        Читает файл .xlsx и выдает информацию о нем в виде списка объектов словаря.

        f   <-  Файл таблицы в формате .xls/.xlsx
        """
        df = pd.read_excel(f, 'Спецификация', skiprows=6)
        for row_index, row in df.iterrows():
            try:
                yield {
                    'index': row_index,
                    'comp_code': row.get('Артикул'),
                    'comp_name': row.get('Наименование/описание'),
                    'comp_unit': row.get('Ед. изм.'),
                    'comp_count': row.get('Кол-во'),
                    'our_code': row.get('Артикул ITK/IEK'),
                    'our_name': row.get('Наименование ITK/IEK'),
                    'our_unit': row.get('Ед. изм.. 1'),
                    'our_count': row.get('Кол-во. 1'),
                    'our_cost': row.get('Цена с НДС, руб.'),
                    'our_cost_amount': row.get('Сумма с НДС, руб.'),
                    'comment': row.get('Комментарии'),
                    'availability_status': row.get('Статус наличия'),
                    'promotion': row.get('Акция'),
                    'analogues': list()
                }
            except ValueError:
                raise ValueError('Невалидное значение: ')

    @staticmethod
    def analog_to_row(analog, row, similar: list = None, ex_m: bool = True):
        """
        Возвращает объект строки в виде словаря для добавления
        в список данных во время пересчета.
        """
        row['our_code'] = analog.code
        row['our_name'] = analog.name
        row['our_unit'] = analog.unit.name
        comp_count = row.get('comp_count')
        if comp_count:
            row['our_count'] = comp_count if comp_count <= analog.count else analog.count
        else:
            row['our_count'] = 0
        row['our_cost'] = analog.cost
        row['our_cost_amount'] = analog.cost * row.get('our_count')
        row['comment'] = analog.comment
        row['availability_status'] = 'В наличии' if analog.count > 0 else 'Нет в наличии'
        row['promotion'] = analog.promotion_description if analog.is_promotion else ''
        row['exactly_match'] = ex_m
        if similar:
            row['similar'] = similar
        return row

    @classmethod
    def search_analog(cls, row: dict, equipment: OurEquipment.objects, prc: SearchProcess):
        """
        Получает строку с данными об оборудовании конкурента и возвращает два значения:
        первично подобранный аналог и список (QuerySet) вторично подобранных аналогов в
        случае успешного поиска.
        В случае неудачного поиска возвращает None и пустой QuerySet.

        row         <-  Объект словаря с данными из строки сметы
        equipment   <-  Список номенклатуры, из которой необходимо подобрать аналог
        prc         <-  Запущенный процесс пересчета
        """
        comp_code = row.get('comp_code')
        comp_name = row.get('comp_name')
        exactly_match = True

        # поиск среди артикулов номенклатуры ИЕК
        if equipment.filter(code=comp_code).exists():
            anlg = equipment.get(code=comp_code)
            prc.results.create(comp_code, comp_name, anlg.code, anlg.name, True, match_in_own_codes=True)
            prc.match_rows_count += 1
            log = f'Найдено: {comp_code} {comp_name} --> {anlg.code} {anlg.name}'
            cls.process_log(prc, log, cls.remtime_str(prc.remtime()))
            LogService.log('success-recalculate-row', f'Артикул: {comp_code} -> {anlg.code}', 'recalculates', prc.user)
            return anlg, equipment.none(), exactly_match

        # поиск среди наименований номенклатуры ИЕК
        if equipment.filter(name=comp_name).exists():
            anlg = equipment.filter(name=comp_name).first()
            prc.results.create(comp_code, comp_name, anlg.code, anlg.name, True, match_in_own_names=True)
            prc.match_rows_count += 1
            log = f'Найдено: {comp_code} {comp_name} --> {anlg.code} {anlg.name}'
            cls.process_log(prc, log, cls.remtime_str(prc.remtime()))
            LogService.log('success-recalculate-row', f'Артикул: {comp_code} -> {anlg.code}', 'recalculates', prc.user)
            return anlg, equipment.none(), exactly_match

        # поиск среди артикулов номенклатуры конкурентов
        if CompetitorsEquipment.objects.filter(code=comp_code).exists():
            comp_analog = CompetitorsEquipment.objects.filter(code=comp_code).first()
            similar = equipment.filter(keywords__comp_equipment=comp_analog)
            if similar.exists():
                anlg = similar.first()
                similar = similar.exclude(id=anlg.id)
                prc.results.create(comp_code, comp_name, anlg.code, anlg.name, True, match_in_comp_codes=True)
                prc.match_rows_count += 1
                log = f'Найдено: {comp_code} {comp_name} --> {anlg.code} {anlg.name}'
                cls.process_log(prc, log, cls.remtime_str(prc.remtime()))
                LogService.log('success-recalculate-row', f'Артикул: {comp_code} -> {anlg.code}', 'recalculates',
                               prc.user)
                return anlg, similar, exactly_match

        # поиск среди наименований номенклатуры конкурентов
        if CompetitorsEquipment.objects.filter(name=comp_name).exists():
            comp_analog = CompetitorsEquipment.objects.filter(name=comp_name).first()
            similar = equipment.filter(keywords__comp_equipment=comp_analog)
            if similar.exists():
                anlg = similar.first()
                similar = similar.exclude(id=anlg.id)
                prc.results.create(comp_code, comp_name, anlg.code, anlg.name, True, match_in_comp_names=True)
                prc.match_rows_count += 1
                log = f'Найдено: {comp_code} {comp_name} --> {anlg.code} {anlg.name}'
                cls.process_log(prc, log, cls.remtime_str(prc.remtime()))
                LogService.log('success-recalculate-row', f'Артикул: {comp_code} -> {anlg.code}', 'recalculates',
                               prc.user)
                return anlg, similar, exactly_match

        # поиск среди ключей по артикулу
        if KeyWord.objects.filter(keyword=comp_code).exists():
            similar = equipment.filter(
                id__in=KeyWord.objects.filter(keyword=comp_code).values_list('our_equipment_id', flat=True)
            )
            if similar.exists():
                anlg = similar.first()
                similar = similar.exclude(id=anlg.id)
                prc.results.create(comp_code, comp_name, anlg.code, anlg.name, True, match_in_keys=True)
                prc.match_rows_count += 1
                log = f'Найдено: {comp_code} {comp_name} --> {anlg.code} {anlg.name}'
                cls.process_log(prc, log, cls.remtime_str(prc.remtime()))
                LogService.log('success-recalculate-row', f'Артикул: {comp_code} -> {anlg.code}', 'recalculates',
                               prc.user)
                return anlg, similar, exactly_match

        # поиск среди ключей по наименованию
        if KeyWord.objects.filter(keyword=comp_name).exists():
            similar = equipment.filter(
                id__in=KeyWord.objects.filter(keyword=comp_name).values_list('our_equipment_id', flat=True)
            )
            if similar.exists():
                anlg = similar.first()
                similar = similar.exclude(id=anlg.id)
                prc.results.create(comp_code, comp_name, anlg.code, anlg.name, True, match_in_keys=True)
                prc.match_rows_count += 1
                log = f'Найдено: {comp_code} {comp_name} --> {anlg.code} {anlg.name}'
                cls.process_log(prc, log, cls.remtime_str(prc.remtime()))
                LogService.log('success-recalculate-row', f'Артикул: {comp_code} -> {anlg.code}', 'recalculates',
                               prc.user)
                return anlg, similar, exactly_match

        # поиск по совпадению характеристик
        else:
            exactly_match = False
            words = comp_name.split(' ')
            all_values = EquipmentCategoryPropertyValue.objects.filter(value__in=words)
            c = collections.Counter(all_values.values_list('property__category_id', flat=True))
            category_id, _ = max(c.items(), key=lambda p: p[::-1])
            equipment = equipment.filter(category_id=category_id)
            values = all_values.filter(property__category_id=category_id)
            property_count = EquipmentCategoryProperty.objects.filter(category_id=category_id).count()
            annotate_equipment = equipment.annotate(
                property_coincidence_count=Count('properties__value', Q(properties__value__in=values)),
                # psp   <-  Property Similarity Percentage
                psp=F('property_coincidence_count') / Value(property_count, IntegerField()) * Value(100, IntegerField())
            )
            similar = annotate_equipment.filter(property_considence_count__gt=0, psp__gte=prc.psp).order_by(
                '-property_coincidence_count')
            if similar.exists():
                anlg = similar.first()
                similar = similar.exclude(id=anlg.id)
                prc.results.create(comp_code, comp_name, anlg.code, anlg.name, True, match_by_properties=True)
                prc.match_rows_count += 1
                log = f'Найдено: {comp_code} {comp_name} --> {anlg.code} {anlg.name}'
                cls.process_log(prc, log, cls.remtime_str(prc.remtime()))
                LogService.log('success-recalculate-row', f'Артикул: {comp_code} -> {anlg.code}', 'recalculates',
                               prc.user)
                return anlg, similar, exactly_match
        prc.results.create(comp_code, comp_name, row.get('code'), row.get('name'), is_unmatch=True)
        prc.unmatch_rows_count += 1
        cls.process_log(prc, f'Не найдено: {comp_code} {comp_name}', cls.remtime_str(prc.remtime()))
        LogService.log('failed-recalculate-row', f'Артикул: {comp_code}; Наименование: {comp_name}', 'recalculates',
                       prc.user)
        return None, equipment.none(), exactly_match

    @classmethod
    def do_recalculate(cls, data: list, prc: SearchProcess, exc_mdls: list):
        """
        Осуществляет поиск по массиву переданных данных, который должен иметь структуру,
        как у выходного списка функции get_prepare_data.

        data    <-  Массив данных в виде списка словарей
        prc     <-  Объект процесса поиска/пересчета
        ex_mdls <-  Список id линеек, которые требуется исключить из подбора
        """

        LogService.log('request-to-recalculate',
                       f'Запрос на пересчет. Пользователь: {prc.user.__str__()}; Email: {prc.user.email};',
                       'recalculates', prc.user)

        recalculated_data = list()
        prc.rows_count = len(data)
        prc.save()
        if prc.rows_count < prc.user.rows_to_recalculate_available:
            prc.error(f'Недостаточно строк. Кол-во строк у пользователя: {prc.user.rows_to_recalculate_available}')
            raise Interrupt('Количество доступных для пересчета строк меньше чем имеется в загруженном файле. '
                            'Поиск остановлен. Обратитесь к администратору')
        our_equipment = OurEquipment.objects.exclude(model__in=exc_mdls)

        for row in data:
            # проверка внешнего прерывания пересчета
            cls.check_interrupt(prc)
            recalculated_row = row
            if row.get('comp_unit') or row.get('comp_count'):
                analog, similar, ex_m = cls.search_analog(row, our_equipment, prc)
                if analog():
                    recalculated_row = cls.analog_to_row(analog, row, ex_m)
            else:
                prc.results.create(row.get('comp_code'), row.get('comp_name'), row.get('our_code'), row.get('our_name'),
                                   is_skip=True)
                prc.skip_rows_count += 1
            recalculated_data.append(recalculated_row)
            prc.save()
        prc.success()  # Завершение процесса пересчета
        return recalculated_data

    @staticmethod
    def get_file(data: list):
        """
        Генерирует стилизованный файл со списком аналогов и возвращает путь до него.

        data    <-  Список строк, которые необходимо занести в итоговый файл пересчета
        """
        wb = Workbook()
        ws = wb.active

        ws.title = 'Спецификация'

        # выставляем ширину столбцов
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 6
        ws.column_dimensions['E'].width = 6
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 6
        ws.column_dimensions['I'].width = 6
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 20
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15

        # добавляем контактные данные
        ws['G1'] = 'Контактный телефон:'
        ws['G2'] = '7 (495) 542-22-22'
        ws['G3'] = 'www.iek.ru'
        ws['G1'].font = document_styles.standard_font
        ws['G2'].font = document_styles.standard_font
        ws['G3'].font = document_styles.standard_font

        # вставка логотипа ИЕК
        logo = Image(os.path.join(settings.BASE_DIR, "iek-logo.png"))
        logo.height = 50
        logo.width = 100
        ws.add_image(logo, 'D1')

        # заголовок столбца номера строки
        num_title = ws['A7']
        num_title.value = '№ п.п.'
        num_title.font = document_styles.title_font
        num_title.fill = document_styles.white_fill
        num_title.border = document_styles.bold_border

        # заголовок столбца артикула конкурента
        comp_code_title = ws['B7']
        comp_code_title.value = 'Артикул'
        comp_code_title.font = document_styles.title_font
        comp_code_title.fill = document_styles.white_fill
        comp_code_title.border = document_styles.bold_border

        # заголовок столбца названия конкурента
        comp_name_title = ws['C7']
        comp_name_title.value = 'Наименование/Описание'
        comp_name_title.font = document_styles.title_font
        comp_name_title.fill = document_styles.white_fill
        comp_name_title.border = document_styles.bold_border

        # заголовок столбца единицы измерения товара у конкурента
        comp_units_title = ws['D7']
        comp_units_title.value = 'Ед. изм.'
        comp_units_title.font = document_styles.title_font
        comp_units_title.fill = document_styles.white_fill
        comp_units_title.border = document_styles.bold_border

        # заголовок столбца количества товара у конкурента
        comp_count_title = ws['E7']
        comp_count_title.value = 'Кол-во'
        comp_count_title.font = document_styles.title_font
        comp_count_title.fill = document_styles.white_fill
        comp_count_title.border = document_styles.bold_border

        # заголовок столбца артикула ИЕК
        code_title = ws['F7']
        code_title.value = 'Артикул ITK/IEK'
        code_title.font = document_styles.title_font
        code_title.fill = document_styles.iek_title_fill
        code_title.border = document_styles.bold_border

        # заголовок столбца названия ИЕК
        name_title = ws['G7']
        name_title.value = 'Наименование ITK/IEK'
        name_title.font = document_styles.title_font
        name_title.fill = document_styles.iek_title_fill
        name_title.border = document_styles.bold_border

        # заголовок столбца единицы измерения товара ИЕК
        units_title = ws['H7']
        units_title.value = 'Ед. изм.'
        units_title.font = document_styles.title_font
        units_title.fill = document_styles.iek_title_fill
        units_title.border = document_styles.bold_border

        # заголовок столбца количества товара ИЕК
        count_title = ws['I7']
        count_title.value = 'Кол-во'
        count_title.font = document_styles.title_font
        count_title.fill = document_styles.iek_title_fill
        count_title.border = document_styles.bold_border

        # заголовок столбца цены с НДС ИЕК
        cost_with_nds_title = ws['J7']
        cost_with_nds_title.value = 'Цена с НДС, руб.'
        cost_with_nds_title.font = document_styles.title_font
        cost_with_nds_title.fill = document_styles.iek_title_fill
        cost_with_nds_title.border = document_styles.bold_border

        # заголовок столбца суммы с НДС ИЕК
        amount_with_nds_title = ws['K7']
        amount_with_nds_title.value = 'Сумма с НДС, руб.'
        amount_with_nds_title.font = document_styles.title_font
        amount_with_nds_title.fill = document_styles.iek_title_fill
        amount_with_nds_title.border = document_styles.bold_border

        # заголовок столбца комментариев
        comments_title = ws['L7']
        comments_title.value = 'Комментарии'
        comments_title.font = document_styles.title_font
        comments_title.fill = document_styles.comment_title_fill
        comments_title.border = document_styles.bold_border

        # заголовок столбца статуса наличия
        availability_status_title = ws['M7']
        availability_status_title.value = 'Статус наличия'
        availability_status_title.font = document_styles.title_font
        availability_status_title.fill = document_styles.comment_title_fill
        availability_status_title.border = document_styles.bold_border

        # заголовок столбца акции
        discount_title = ws['N7']
        discount_title.value = 'Акция'
        discount_title.font = document_styles.title_font
        discount_title.fill = document_styles.discount_title_fill
        discount_title.border = document_styles.bold_border

        row_num = 1
        row_start = 7

        for data_row in data:
            cell_num_str = str(row_num + row_start)
            num = ws['A' + cell_num_str]
            comp_code = ws['B' + cell_num_str]
            comp_name = ws['C' + cell_num_str]
            comp_units = ws['D' + cell_num_str]
            comp_count = ws['E' + cell_num_str]
            code = ws['F' + cell_num_str]
            name = ws['G' + cell_num_str]
            units = ws['H' + cell_num_str]
            count = ws['I' + cell_num_str]
            cost_with_nds = ws['J' + cell_num_str]
            amount_with_nds = ws['K' + cell_num_str]
            comments = ws['L' + cell_num_str]
            availability_status = ws['M' + cell_num_str]
            discount = ws['N' + cell_num_str]

            num.font = document_styles.standard_font
            comp_code.font = document_styles.standard_font
            comp_name.font = document_styles.standard_font
            comp_units.font = document_styles.standard_font
            comp_count.font = document_styles.standard_font
            code.font = document_styles.standard_font
            name.font = document_styles.standard_font
            units.font = document_styles.standard_font
            count.font = document_styles.standard_font
            cost_with_nds.font = document_styles.standard_font
            amount_with_nds.font = document_styles.standard_font
            comments.font = document_styles.standard_font
            availability_status.font = document_styles.standard_font
            discount.font = document_styles.standard_font

            num.fill = document_styles.white_fill
            comp_code.fill = document_styles.white_fill
            comp_name.fill = document_styles.white_fill
            comp_units.fill = document_styles.white_fill
            comp_count.fill = document_styles.white_fill
            code.fill = document_styles.iek_cell_fill
            name.fill = document_styles.iek_cell_fill
            units.fill = document_styles.iek_cell_fill
            count.fill = document_styles.iek_cell_fill
            cost_with_nds.fill = document_styles.iek_cell_fill
            amount_with_nds.fill = document_styles.iek_cell_fill
            comments.fill = document_styles.comment_cell_fill
            availability_status.fill = document_styles.comment_cell_fill
            discount.fill = document_styles.discount_cell_fill

            num.border = document_styles.standard_border
            comp_code.border = document_styles.standard_border
            comp_name.border = document_styles.standard_border
            comp_units.border = document_styles.standard_border
            comp_count.border = document_styles.standard_border
            code.border = document_styles.separate_left_bold_border
            name.border = document_styles.standard_border
            units.border = document_styles.standard_border
            count.border = document_styles.standard_border
            cost_with_nds.border = document_styles.standard_border
            amount_with_nds.border = document_styles.standard_border
            comments.border = document_styles.separate_left_bold_border
            availability_status.border = document_styles.standard_border
            discount.border = document_styles.separate_left_bold_border

            num.alignment = document_styles.alignment_center
            comp_code.alignment = document_styles.alignment_left
            comp_name.alignment = document_styles.alignment_left
            comp_units.alignment = document_styles.alignment_center
            comp_count.alignment = document_styles.alignment_center
            code.alignment = document_styles.alignment_left
            name.alignment = document_styles.alignment_left
            units.alignment = document_styles.alignment_center
            count.alignment = document_styles.alignment_center
            cost_with_nds.alignment = document_styles.alignment_center
            amount_with_nds.alignment = document_styles.alignment_center
            comments.alignment = document_styles.alignment_left
            availability_status.alignment = document_styles.alignment_left
            discount.alignment = document_styles.alignment_left

            num.value = str(row_num)
            comp_code.value = data_row['comp_code']
            comp_name.value = data_row['comp_name']
            code.value = data_row['code']
            name.value = data_row['name']

            row_num += 1
        now_str = timezone.localtime(timezone.now()).strftime('%Y-%m-%d_%H-%M-%S')
        filepath = os.path.join(settings.MEDIA_ROOT, 'tmp_recalculates', f'recalculate_{now_str}.xlsx')
        wb.save(filename=filepath)
        return filepath

    @staticmethod
    def get_file_response(path: str, email: str, rm: int = None):
        now = timezone.localtime(timezone.now() + timezone.timedelta(hours=3))
        output_filename = 'Пересчет ASIST ' + email + ' ' + now.strftime("%d.%m.%Y %Hч %Mм") + '.xlsx'
        response = FileResponse(open(path, 'rb'), as_attachment=True, filename=output_filename)
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response['Content-Disposition'] = "attachment; filename=*=UTF-8''" + escape_uri_path(output_filename)
        if rm:
            # удаление временного файла через 'rm' секунд
            threading.Timer(rm, lambda p: os.remove(p), args=[path]).start()
        return response


class UploadDatafileService:
    nomenclature_iek_require_columns = {
        'Артикул': 'code',
        'Наименование': 'name',
        'Категория': 'category',
        'Линейка': 'model_class',
        'Ед. изм.': 'unit',
        'Остаток': 'warehouse balance',
        'Цена с НДС': 'cost',
        'Комментарии': 'comment',
        'Акции': 'promotion_description',
        'Фото': 'photo_list',
    }
    nomenclature_competitor_require_columns = {
        'Компания': 'competitor',
        'Артикул': 'code',
        'Наименование': 'name',
        'Категория': 'category',
        'Ед. изм.': 'unit',
    }
    nomenclature_keys_require_columns = {
        'Артикул ITK/IEK': 'code',
        'Артикул конкурента': 'comp_code',
        'Поисковой ключ': 'search_key_string',
        'Тип ключа': 'search_key_type'
    }
    category_require_columns = {
        'Категория': 'category',
        'Свойство': 'property',
        'Значение': 'value'
    }
    key_type_signs = {
        'N': 'name',
        'C': 'code',
        'B': 'both',
    }

    @staticmethod
    def get_row_data(row_items, model_columns):
        """
            Возвращает словарь с данными из строки файла Excel,
            содержащий информацию о номенклатуре и ее характеристиках
        """
        data = dict()
        data['properties'] = dict()
        for column_index, value in row_items:  # получение сырых данных из строки
            try:
                data[model_columns[column_index]] = value
            except KeyError:  # если названия колонки нет в списке обязательных, значит это характеристика
                data['properties'][column_index] = value
        return data

    @staticmethod
    def check_required_columns(data, columns):
        """
            Проверяет у словаря с данными о номенклатуре наличие всех
            необходимых атрибутов по ключам
        """
        for k, v in columns.items():  # проверка на наличие всех обязательных колонок
            try:
                data[v]
            except KeyError:
                raise InvalidUploadingData(
                    f'Не хватает колонки "{k}". Исправьте ваш шаблонный файл или скачайте его заново')

    @staticmethod
    def fetch_new_equipment_photo(nom, url_list):
        for img in nom.images.all():
            img.img.delete(save=False)
            img.delete()
        for url in url_list:
            try:
                name, _ = urlretrieve(url)
                new_img = OurEquipmentImage(equipment_id=nom.id)
                new_img.img.save("nomphoto.jpg", File(open(name, 'rb')))
                new_img.save()
            except (URLError, ValueError):
                urlcleanup()
            finally:
                urlcleanup()

    @staticmethod
    def set_our_equipment_properties(equipment: OurEquipment, properties: dict, category_name: str):
        for prop, value in properties.items():
            try:
                category = EquipmentCategory.objects.get(name=category_name)
                cat_prop = EquipmentCategoryProperty.objects.get(name=prop, category=category)
                cat_prop_val, cpv_created = EquipmentCategoryPropertyValue.objects.get_or_create(property=cat_prop,
                                                                                                 value=value)
                nom_prop, np_created = OurEquipmentProperty.objects.get_or_create(
                    property=cat_prop, equipment=equipment, value=cat_prop_val
                )
                if not cpv_created:
                    cat_prop_val.save()
                if not np_created:
                    nom_prop.save()
            except EquipmentCategoryProperty.DoesNotExist:
                raise InvalidUploadingData(f'У категории {category_name} нет свойства {prop}. '
                                           f'Добавьте свойство через панель DBA или удалите его '
                                           f'из загружаемого файла.')
            except IntegrityError:
                raise InvalidUploadingData(f'Ошибка повторения уникальных значений')

    @staticmethod
    def set_comp_equipment_properties(equipment: CompetitorsEquipment, properties: dict, category_name: str):
        for prop, value in properties.items():
            try:
                category = EquipmentCategory.objects.get(name=category_name)
                cat_prop = EquipmentCategoryProperty.objects.get(name=prop, category=category)
                cat_prop_val, cpv_created = EquipmentCategoryPropertyValue.objects.get_or_create(property=cat_prop,
                                                                                                 value=value)
                if CompetitorsEquipmentProperty.objects.filter(property=cat_prop, equipment=equipment).exists():
                    nom_prop = CompetitorsEquipmentProperty.objects.get(property=cat_prop, equipment=equipment)
                    nom_prop.value = cat_prop_val
                else:
                    nom_prop = CompetitorsEquipmentProperty(property=cat_prop, equipment=equipment, value=cat_prop_val)
                nom_prop.save()
                if not cpv_created:
                    cat_prop_val.save()
            except EquipmentCategoryProperty.DoesNotExist:
                raise InvalidUploadingData(f'У категории {category_name} нет свойства {prop}. '
                                           f'Добавьте свойство через панель DBA или удалите его '
                                           f'из загружаемого файла.')
            except IntegrityError:
                raise InvalidUploadingData(f'Ошибка повторения уникальных значений')

    @classmethod
    @transaction.atomic
    def upload_our_equipment(cls, file):
        """
            Принимает файл Excel соответствующий шаблону, вытаскивает из него
            данные об оборудовании ITK/IEK и загружает данные в базу данных.
            Если номенклатура была в базе данных, ее данные обновятся, если нет,
            она будет создана. Если данные в строке не позволяют сохранить их
            в базу данных, пропускает строку и увеличивает кол-во непрошедших
            валидацию строк. Возвращает кол-во непрошедших валидацию строк.
        """
        invalid = 0  # кол-во пропущенных строк, не прошедших валидацию
        valid = 0  # кол-во строк, прошедших валидацию
        df = pd.read_excel(file, sheet_name='DATA')
        for row_index, row in df.iterrows():
            data = cls.get_row_data(row.items(), cls.nomenclature_iek_require_columns)
            cls.check_required_columns(data, cls.nomenclature_iek_require_columns)

            code = data.get('code')
            name = data.get('name')
            category = data.get('category')
            cost = data.get('cost')
            comment = data.get('comment')
            promotion_description = data.get('promotion_description')
            photo_list = data.get('photo_list')
            photo_list = photo_list.split(' ') if not pd.isna(photo_list) else list()

            # если в строке не указан артикул или наименование, пропускаем строку как невалидную
            if pd.isna(code) or pd.isna(name):
                invalid += 1
                continue

            # если указана категория номенклатуры, которой нет в базе данных, пропускаем строку как невалидную
            if not pd.isna(category) and not EquipmentCategory.objects.filter(name=category).exists():
                invalid += 1
                continue

            try:
                equipment = OurEquipment.objects.get(code=data.get('code'))
            except OurEquipment.DoesNotExist:
                equipment = OurEquipment(code=code)

            equipment.approved = False
            equipment.name = name
            equipment.category = EquipmentCategory.objects.get(name=category)
            if not pd.isna(cost):
                equipment.cost = cost
            if not pd.isna(comment):
                equipment.comment = str(comment)
            if not pd.isna(promotion_description):
                equipment.is_promotion = True
                equipment.promotion_description = str(promotion_description)
            else:
                equipment.is_promotion = False
                equipment.promotion_description = ''

            equipment.save()

            if len(photo_list):
                cls.fetch_new_equipment_photo(equipment, photo_list)
            cls.set_our_equipment_properties(equipment, data.get('properties'), category)

            valid += 1

        return invalid, valid

    @classmethod
    def upload_comp_equipment(cls, file):
        """
            Принимает файл Excel соответствующий шаблону, вытаскивает из него
            данные об оборудовании конкурентов и загружает данные в базу данных.
            Если номенклатура была в базе данных, ее данные обновятся, если нет,
            она будет создана. Если данные в строке не позволяют сохранить их
            в базу данных, пропускает строку и увеличивает кол-во непрошедших
            валидацию строк. Возвращает кол-во непрошедших валидацию строк.
        """
        invalid = 0  # кол-во пропущенных строк, не прошедших валидацию
        valid = 0  # кол-во строк, прошедших валидацию
        df = pd.read_excel(file, sheet_name='DATA')
        for row_index, row in df.iterrows():
            data = cls.get_row_data(row.items(), cls.nomenclature_competitor_require_columns)
            cls.check_required_columns(data, cls.nomenclature_competitor_require_columns)

            competitor = data.get('competitor')
            code = data.get('code')
            name = data.get('name')
            category = data.get('category')

            # если в строке не указан артикул или наименование, пропускаем строку как невалидную
            if pd.isna(code):
                invalid += 1
                continue
            if pd.isna(name):
                name = None

            # если указана категория номенклатуры, которой нет в базе данных, пропускаем строку как невалидную
            if not pd.isna(category) and not EquipmentCategory.objects.filter(name=category).exists():
                invalid += 1
                continue

            try:
                nom = CompetitorsEquipment.objects.get(code=data.get('code'))
            except CompetitorsEquipment.DoesNotExist:
                nom = CompetitorsEquipment(code=code)

            nom.name = name
            nom.competitor, _ = Competitor.objects.get_or_create(name=competitor)
            nom.category = EquipmentCategory.objects.get(name=category)

            nom.save()
            cls.set_comp_equipment_properties(nom, data.get('properties'), category)

            valid += 1

        return invalid, valid

    @classmethod
    def upload_keywords(cls, file):
        """
            Принимает файл Excel соответствующий шаблону, вытаскивает из него
            данные о связи номенклатуры ITK/IEK и номенклатуры конурентов
            и загружает данные в базу данных.
            Если данные в строке не позволяют сохранить их
            в базу данных, пропускает строку и увеличивает кол-во непрошедших
            валидацию строк. Возвращает кол-во непрошедших валидацию строк.
        """
        invalid = 0  # кол-во пропущенных строк, не прошедших валидацию
        valid = 0  # кол-во строк, прошедших валидацию
        df = pd.read_excel(file, sheet_name='DATA')
        for row_index, row in df.iterrows():
            data = cls.get_row_data(row.items(), cls.nomenclature_keys_require_columns)
            cls.check_required_columns(data, cls.nomenclature_keys_require_columns)

            code = data.get('code')
            comp_code = data.get('comp_code')
            search_key_string = data.get('search_key_string')
            search_key_type = data.get('search_key_type')

            # если в одном из столбцов нет данных, строка считается невалидной
            if pd.isna(code) or pd.isna(comp_code) or pd.isna(search_key_string) or pd.isna(search_key_type):
                invalid += 1
                continue

            try:
                our_equipment = OurEquipment.objects.get(code=code)
                comp_equipment = CompetitorsEquipment.objects.get(code=comp_code)
            except (OurEquipment.DoesNotExist, CompetitorsEquipment.DoesNotExist,):
                invalid += 1
                continue

            try:
                kw = KeyWord.objects.get(
                    keyword=search_key_string, our_equipment=our_equipment, comp_equipment=comp_equipment)
                kw.approved = False
                kw.type = cls.key_type_signs.get(search_key_type)
                kw.save()
            except KeyWord.DoesNotExist:
                kw = KeyWord(keyword=search_key_string, our_equipment=our_equipment, comp_equipment=comp_equipment,
                             is_approved=False)
                kw.save()
            valid += 1
        return invalid, valid

    @classmethod
    def upload_equipment_categories(cls, file):
        """
            Принимает файл Excel соответствующий шаблону, вытаскивает из него
            данные о связи номенклатуры ITK/IEK и номенклатуры конурентов
            и загружает данные в базу данных.
            Если данные в строке не позволяют сохранить их
            в базу данных, пропускает строку и увеличивает кол-во непрошедших
            валидацию строк. Возвращает кол-во непрошедших валидацию строк.
        """
        invalid = 0  # кол-во пропущенных строк, не прошедших валидацию
        valid = 0  # кол-во строк, прошедших валидацию
        df = pd.read_excel(file, sheet_name='DATA')
        for row_index, row in df.iterrows():
            data = cls.get_row_data(row.items(), cls.category_require_columns)
            cls.check_required_columns(data, cls.category_require_columns)

            category_name = data.get('category')
            property_name = data.get('property')
            value_str = data.get('value')
            try:
                category, _ = EquipmentCategory.objects.get_or_create(name=category_name)
                property, _ = EquipmentCategoryProperty.objects.get_or_create(category=category, name=property_name)
                value, _ = EquipmentCategoryPropertyValue.objects.get_or_create(property=property, value=value_str)
                valid += 1
            except:
                invalid += 1

        return invalid, valid


class DownloadDatafileService:
    @staticmethod
    def create_kwds_datafile(kwds):
        df = pd.DataFrame([pd.Series({
            'Артикул ITK/IEK': d.get('our_equipment__code'),
            'Артикул конкурента': d.get('comp_equipment__code'),
            'Поисковой ключ': d.get('keyword'),
        }) for d in kwds.order_by('our_equipment').values('our_equipment__code', 'comp_equipment__code', 'keyword')])
        ltime = timezone.localtime(timezone.now()).strftime('%Y-%m-%d_%H-%M')
        fld = os.path.join(settings.MEDIA_ROOT, 'download', 'kwds')
        if not os.path.exists(fld):
            os.makedirs(fld)
        path = os.path.join(fld, f'asist_kwds_{ltime}.xlsx')
        df.to_excel(path, index=False, sheet_name='DATA')
        return path

    @staticmethod
    def create_our_equipment_datafile(equipment, wop: bool = False):
        """
        wop -   With Out Properties. При значении True не выгружает в файл свойства номенклатуры.
                Позволяет скачивать номенклатуры нескольких категорий в одном файле.
        """
        if not wop and equipment.order_by('category_id').distinct('category_id').count() > 1:
            raise MultiEquipCategoryDownload
        df = pd.DataFrame([pd.Series({
            'Артикул': d.get('code'),
            'Наименование': d.get('name'),
            'Категория': d.get('category__name'),
            'Линейка': d.get('model__name'),
            'Ед. изм': d.get('unit__name'),
            'Остаток': d.get('count'),
            'Цена с НДС': d.get('cost'),
            'Комментарии': d.get('description'),
            'Акции': d.get('sale_description') if d.get('is_sale') else '',
            'Фото': '',
            # разворачивание списка свойств номенклатуры в виде столбцов в выходном файле
            **{p.get('property__name'): p.get('value__value')
               for p in OurEquipmentProperty.objects.filter(
                    property__category_id=d.get('category_id'), equipment_id=d.get('id')
                ).values('property__name', 'value__value') if not wop}
        }) for d in equipment.values(
            'id', 'category_id', 'code', 'name', 'category__name', 'model__name', 'unit__name', 'count', 'cost',
            'description', 'sale_description', 'is_sale'
        )])
        ltime = timezone.localtime(timezone.now()).strftime('%Y-%m-%d_%H-%M')
        fld = os.path.join(settings.MEDIA_ROOT, 'download', 'our_equipment')
        if not os.path.exists(fld):
            os.makedirs(fld)
        path = os.path.join(fld, f'asist_our_equipment_{ltime}.xlsx')
        df.to_excel(path, index=False, sheet_name='DATA')
        return path

    @staticmethod
    def create_comp_equipment_datafile(equipment, wop: bool = False):
        if not wop and equipment.order_by('category_id').distinct('category_id').count() > 1:
            raise MultiEquipCategoryDownload
        df = pd.DataFrame([pd.Series({
            'Компания': d.get('competitor__name'),
            'Артикул': d.get('code'),
            'Наименование': d.get('name'),
            'Категория': d.get('category__name'),
            'Ед. изм': d.get('unit__name'),
            # разворачивание списка свойств номенклатуры конкурента в виде столбцов в выходном файле
            **{p.get('property__name'): p.get('value__value')
               for p in CompetitorsEquipmentProperty.objects.filter(
                    property__category_id=d.get('category_id'), equipment_id=d.get('id')
                ).values('property__name', 'value__value') if not wop}
        }) for d in equipment.values(
            'id', 'category_id', 'competitor__name', 'code', 'name', 'category__name')])
        ltime = timezone.localtime(timezone.now()).strftime('%Y-%m-%d_%H-%M')
        fld = os.path.join(settings.MEDIA_ROOT, 'download', 'comp_equipment')
        if not os.path.exists(fld):
            os.makedirs(fld)
        path = os.path.join(fld, f'asist_comp_equipment_{ltime}.xlsx')
        df.to_excel(path, index=False, sheet_name='DATA')
        return path

    @staticmethod
    def create_category_datafile(categories):
        properties_list = list()
        for cat in categories.values('id', 'name'):
            for prop in EquipmentCategoryPropertyValue.objects.filter(property__category_id=cat.get('id')).values(
                    'property__name', 'value'):
                properties_list.append(pd.Series({
                    'Категория': cat.get('name'), 'Свойство': prop.get('property__name'), 'Значение': prop.get('value')
                }))
        df = pd.DataFrame(properties_list).sort_values(['Категория', 'Свойство'])
        ltime = timezone.localtime(timezone.now()).strftime('%Y-%m-%d_%H-%M')
        fld = os.path.join(settings.MEDIA_ROOT, 'downloads', 'categories')
        if not os.path.exists(fld):
            os.makedirs(fld)
        path = os.path.join(fld, f'asist_categories_{ltime}.xlsx')
        df.to_excel(path, index=False, sheet_name='DATA')
        return path

    @staticmethod
    def get_file_response(path: str, filename: str, rm: int = None):
        response = FileResponse(open(path, 'rb'), as_attachment=True, filename=filename)
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response['Content-Disposition'] = "attachment; filename=*=UTF-8''" + escape_uri_path(filename)
        if rm:
            # удаление временного файла через 'rm' секунд
            threading.Timer(rm, lambda p: os.remove(p), args=[path]).start()
        return response
